import streamlit as st
import openpyxl
from collections import defaultdict
from datetime import datetime
import io
import zipfile
import pandas as pd
import plotly.express as px
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import base64

# ========= SEU CÓDIGO ORIGINAL (NÃO ALTERADO) =========
def to_float(valor):
    if valor is None or valor == '': return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    try:
        s = str(valor).strip().replace('.', '').replace(',', '.')
        return float(s)
    except: return 0.0

def extrair_dados_planilha(arquivo_xlsx):
    wb = openpyxl.load_workbook(arquivo_xlsx)
    ws = wb.active
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    notas = defaultdict(lambda: {'serie': None,'numero': None,'dataEmissao': None,'cnpjEmitente': None,'itens': []})
    emitente_info = {}
    for row in range(2, ws.max_row + 1):
        row_data = {h: ws.cell(row=row, column=col).value for col, h in enumerate(headers, 1)}
        chave = row_data.get('ChaveAcesso')
        if chave and row_data.get('DescricaoProduto'):
            if not emitente_info:
                emitente_info = {
                    'cnpj': str(row_data.get('CnpjEmitente') or ''),
                    'razao': str(row_data.get('RazaoSocialEmitente') or row_data.get('NomeEmitente') or 'Não informado'),
                    'ie': str(row_data.get('InscricaoEstadualEmitente') or ''),
                }
            notas[chave]['serie'] = row_data.get('SerieDocumento')
            notas[chave]['numero'] = row_data.get('NumeroDocumento')
            notas[chave]['dataEmissao'] = row_data.get('DataEmissaoNfe')
            notas[chave]['cnpjEmitente'] = row_data.get('CnpjEmitente')
            notas[chave]['itens'].append({
                'codigo': row_data.get('Produto'),
                'descricao': row_data.get('DescricaoProduto'),
                'ncm': row_data.get('NcmProduto'),
                'cfop': row_data.get('CfopProduto'),
                'quantidade': to_float(row_data.get('QuantidadeUnidadeComercial')),
                'valorUnitario': to_float(row_data.get('ValorUnitarioComercial')),
                'valorTotal': to_float(row_data.get('ValorTotalProduto')),
                'valorFrete': to_float(row_data.get('ValorFrete')),
                'valorSeguro': to_float(row_data.get('ValorSeguro')),
                'valorDesconto': to_float(row_data.get('ValorDesconto')),
                'valorOutras': to_float(row_data.get('ValorOutrasDespesas')),
                'valorIcmsBc': to_float(row_data.get('ValorIcmsBc')), # só para relatório
                'valorIcms': to_float(row_data.get('ValorIcms')), # só para relatório
                'icmsTag': row_data.get('TipoIcmsTag'),
                'icmsTributacao': row_data.get('IcmsTributacao'),
            })
    return notas, emitente_info

def gerar_xml(chave, nota):
    data = nota['dataEmissao']
    if isinstance(data, datetime):
        data_str = data.strftime('%Y-%m-%d')
    else:
        data_str = str(data)[:10]
    xml_str = f'''<?xml version="1.0" encoding="utf-8"?><nfeProc xmlns="http://www.portalfiscal.inf.br/nfe" versao="4.00"><NFe xmlns="http://www.portalfiscal.inf.br/nfe"><infNFe Id="NFe{chave}" versao="4.00"><ide><cUF>42</cUF><cNF>{str(nota['numero']).zfill(8)}</cNF><natOp>VENDA DE MERCADORIAS</natOp><mod>65</mod><serie>{nota['serie']}</serie><nNF>{nota['numero']}</nNF><dhEmi>{data_str}T20:10:12-03:00</dhEmi><tpNF>1</tpNF><idDest>1</idDest><cMunFG>4209102</cMunFG><tpImp>4</tpImp><tpEmis>1</tpEmis><cDV>6</cDV><tpAmb>1</tpAmb><finNFe>1</finNFe><indFinal>1</indFinal><indPres>1</indPres><procEmi>0</procEmi><verProc>1.00</verProc></ide><emit><CNPJ>{nota['cnpjEmitente']}</CNPJ><xNome>DNA RESTAURANTE LTDA</xNome><xFant>DNA SUSHI</xFant><enderEmit><xLgr>Rua Monsenhor Gercino</xLgr><nro>1285</nro><xBairro>Itaum</xBairro><cMun>4209102</cMun><xMun>Joinville</xMun><UF>SC</UF><CEP>89210009</CEP><cPais>1058</cPais><xPais>BRASIL</xPais><fone>47991986628</fone></enderEmit><IE>262605775</IE><CRT>1</CRT></emit>'''
    total_prod = 0; total_frete = 0; total_seg = 0; total_desc = 0; total_outro = 0
    for idx, item in enumerate(nota['itens'], 1):
        total_prod += item['valorTotal']; total_frete += item.get('valorFrete', 0); total_seg += item.get('valorSeguro', 0); total_desc += item.get('valorDesconto', 0); total_outro += item.get('valorOutras', 0)
        icms_tag = item.get('icmsTag', 'ICMSSN500')
        if icms_tag == 'ICMSSN102':
            icms_xml = f'''<ICMS><ICMSSN102><orig>0</orig><CSOSN>102</CSOSN></ICMSSN102></ICMS>'''
            pis_xml = f'''<PIS><PISOutr><CST>49</CST><qBCProd>0.0000</qBCProd><vAliqProd>0.0000</vAliqProd><vPIS>0.00</vPIS></PISOutr></PIS>'''
            cofins_xml = f'''<COFINS><COFINSOutr><CST>49</CST><vBC>0.00</vBC><pCOFINS>0.00</pCOFINS><vCOFINS>0.00</vCOFINS></COFINSOutr></COFINS>'''
        else:
            icms_xml = f'''<ICMS><ICMSSN500><orig>0</orig><CSOSN>500</CSOSN><vBCSTRet>0.00</vBCSTRet><pST>0.00</pST><vICMSSTRet>0.00</vICMSSTRet></ICMSSN500></ICMS>'''
            pis_xml = f'''<PIS><PISNT><CST>04</CST></PISNT></PIS>'''
            cofins_xml = f'''<COFINS><COFINSNT><CST>04</CST></COFINSNT></COFINS>'''
        prod_xml = f'''<prod><cProd>{item['codigo']}</cProd><cEAN>SEM GTIN</cEAN><xProd>{item['descricao']}</xProd><NCM>{item['ncm']}</NCM><CFOP>{item['cfop']}</CFOP><uCom>UN</uCom><qCom>{item['quantidade']:.2f}</qCom><vUnCom>{item['valorUnitario']:.2f}</vUnCom><vProd>{item['valorTotal']:.2f}</vProd><cEANTrib>SEM GTIN</cEANTrib><uTrib>UN</uTrib><qTrib>{item['quantidade']:.2f}</qTrib><vUnTrib>{item['valorUnitario']:.2f}</vUnTrib>'''
        if item.get('valorFrete', 0) > 0: prod_xml += f'''<vFrete>{item['valorFrete']:.2f}</vFrete>'''
        if item.get('valorSeguro', 0) > 0: prod_xml += f'''<vSeg>{item['valorSeguro']:.2f}</vSeg>'''
        if item.get('valorDesconto', 0) > 0: prod_xml += f'''<vDesc>{item['valorDesconto']:.2f}</vDesc>'''
        if item.get('valorOutras', 0) > 0: prod_xml += f'''<vOutro>{item['valorOutras']:.2f}</vOutro>'''
        prod_xml += f'''<indTot>1</indTot></prod>'''
        xml_str += f'''<det nItem="{idx}">{prod_xml}<imposto>{icms_xml}{pis_xml}{cofins_xml}</imposto></det>'''
    vNF = total_prod + total_frete + total_seg + total_outro - total_desc
    xml_str += f'''<total><ICMSTot><vBC>0.00</vBC><vICMS>0.00</vICMS><vICMSDeson>0.00</vICMSDeson><vFCP>0.00</vFCP><vBCST>0.00</vBCST><vST>0.00</vST><vFCPST>0.00</vFCPST><vFCPSTRet>0.00</vFCPSTRet><vProd>{total_prod:.2f}</vProd><vFrete>{total_frete:.2f}</vFrete><vSeg>{total_seg:.2f}</vSeg><vDesc>{total_desc:.2f}</vDesc><vII>0.00</vII><vIPI>0.00</vIPI><vIPIDevol>0.00</vIPIDevol><vPIS>0.00</vPIS><vCOFINS>0.00</vCOFINS><vOutro>{total_outro:.2f}</vOutro><vNF>{vNF:.2f}</vNF></ICMSTot></total><transp><modFrete>9</modFrete></transp><pag><detPag><tPag>04</tPag><vPag>{vNF:.2f}</vPag><card><tpIntegra>2</tpIntegra></card></detPag></pag></infNFe></NFe></nfeProc>'''
    return xml_str
# ========= FIM =========

# --- INTERFACE ORIGINAL ---
st.set_page_config(page_title="Conversor NFe 2.0 Pro", page_icon="📄", layout="wide")
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
html,body{font-family:'Inter',sans-serif}
.hero{background:linear-gradient(135deg,#1e3a8a,#3b82f6);padding:2rem;border-radius:16px;color:white;margin-bottom:1.5rem}
.metric-card{background:white;padding:1rem;border-radius:12px;border:1px solid #e5e7eb;text-align:center}
.metric-value{font-size:1.6rem;font-weight:700;color:#1e3a8a}
div.stDownloadButton>button{background:#1e3a8a;color:white;border-radius:10px;font-weight:600;padding:0.8rem;width:100%}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="hero"><h1>📄 Conversor NFe 2.0 Pro</h1><p>XML + Excel + PDF + Gráfico — Emitente automático do Excel</p></div>', unsafe_allow_html=True)

up = st.file_uploader("Planilha.xlsx", type="xlsx")
if up:
    notas, emitente = extrair_dados_planilha(up)
    itens = [i for n in notas.values() for i in n['itens']]
    df = pd.DataFrame(itens)
    total_bruto = df['valorTotal'].sum(); total_desc = df['valorDesconto'].sum(); total_liq = total_bruto - total_desc
    datas = [n['dataEmissao'] for n in notas.values() if n['dataEmissao']]
    datas_dt = pd.to_datetime(datas, errors='coerce').dropna()
    periodo = f"{datas_dt.min():%d/%m/%Y} a {datas_dt.max():%d/%m/%Y}" if not datas_dt.empty else "-"
    resumo = df.groupby('cfop').agg(Qtde_Itens=('valorTotal','count'),Valor_Bruto=('valorTotal','sum'),Descontos=('valorDesconto','sum')).reset_index()
    resumo['Valor_Liquido'] = resumo['Valor_Bruto'] - resumo['Descontos']

    st.info(f"**Emitente detectado:** {emitente['razao']} • **CNPJ:** {emitente['cnpj']} • **IE:** {emitente['ie']} • **Período:** {periodo}", icon="🏢")
    c1,c2,c3,c4 = st.columns(4)
    for col,lab,val in zip([c1,c2,c3,c4], ["Notas Emitidas","Total Bruto","Descontos","Total Líquido"], [len(notas), total_bruto, total_desc, total_liq]):
        val_str = f"R$ {val:,.2f}" if lab!="Notas Emitidas" else str(int(val))
        col.markdown(f'<div class="metric-card"><div>{lab}</div><div class="metric-value">{val_str}</div></div>', unsafe_allow_html=True)

    colA,colB = st.columns([1.2,1])
    with colA:
        st.subheader("Resumo por CFOP")
        st.dataframe(resumo.style.format({'Valor_Bruto':'R$ {:,.2f}','Descontos':'R$ {:,.2f}','Valor_Liquido':'R$ {:,.2f}'}), use_container_width=True)
    with colB:
        st.subheader("Gráfico por CFOP")
        fig = px.pie(resumo, names='cfop', values='Valor_Liquido', hole=0.4, title="Distribuição Líquida")
        fig.update_traces(textinfo='percent+label')
        st.plotly_chart(fig, use_container_width=True)

    # --- Excel com a nova aba Completo ---
    completo = []
    for chave, nota in notas.items():
        for it in nota['itens']:
            completo.append({
                'Chave': chave, 'Numero': nota['numero'], 'Serie': nota['serie'], 'Data': nota['dataEmissao'],
                'CFOP': it['cfop'], 'Produto': it['descricao'], 'Qtd': it['quantidade'], 'V.Un': it['valorUnitario'],
                'Bruto': it['valorTotal'], 'Desc': it['valorDesconto'], 'Frete': it['valorFrete'],
                'Seg': it['valorSeguro'], 'Outras': it['valorOutras'],
                'Liquido': it['valorTotal']+it['valorFrete']+it['valorSeguro']+it['valorOutras']-it['valorDesconto'],
                'BC ICMS': it['valorIcmsBc'], 'ICMS': it['valorIcms'], 'CST': it['icmsTributacao']
            })
    df_completo = pd.DataFrame(completo)

    excel_buf = io.BytesIO()
    with pd.ExcelWriter(excel_buf, engine='openpyxl') as writer:
        pd.DataFrame([{'Emitente':emitente['razao'],'CNPJ':emitente['cnpj'],'Período':periodo,'Bruto':total_bruto,'Descontos':total_desc,'Líquido':total_liq}]).to_excel(writer, sheet_name='Resumo', index=False)
        resumo.to_excel(writer, sheet_name='Por CFOP', index=False)
        df_completo.to_excel(writer, sheet_name='Completo', index=False)

    pdf_buf = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buf, pagesize=A4)
    styles = getSampleStyleSheet(); elems = [Paragraph(f"<b>Resumo Fiscal</b><br/>{emitente['razao']} - {emitente['cnpj']}", styles['Title'])]
    data = [['CFOP','Itens','Bruto','Desc.','Líquido']]
    for _,r in resumo.iterrows(): data.append([r['cfop'],int(r['Qtde_Itens']),f"R$ {r['Valor_Bruto']:,.2f}",f"R$ {r['Descontos']:,.2f}",f"R$ {r['Valor_Liquido']:,.2f}"])
    t = Table(data); t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e3a8a')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.5,colors.grey)]))
    elems.append(t); doc.build(elems); pdf_buf.seek(0)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf,'w',zipfile.ZIP_DEFLATED) as z:
        for chave,nota in notas.items():
            z.writestr(f"xmls/NFe_{nota['numero']}.xml", gerar_xml(chave, nota))
        z.writestr("resumo/Resumo_NFe.xlsx", excel_buf.getvalue())
        z.writestr("resumo/Resumo_NFe.pdf", pdf_buf.getvalue())
    zip_buf.seek(0)

    st.download_button("⬇️ BAIXAR PACOTE COMPLETO", zip_buf, f"Pacote_NFe_{datetime.now().strftime('%Y%m%d')}.zip", "application/zip")
else:
    st.info("Envie a planilha. O emitente será lido automaticamente do arquivo.")
